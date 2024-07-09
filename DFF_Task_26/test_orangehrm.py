import pytest
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from openpyxl import load_workbook
from webdriver_manager.chrome import ChromeDriverManager
from login_orangehrm import web_page
@pytest.fixture()
def driver():
    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
    driver.implicitly_wait(10)
    yield driver
    driver.close()
    driver.quit()

def read_excel_data(file_path=r"C:\Users\sarit\OneDrive\Documents\user_input_file.xlsx",sheet_name='Sheet1'):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    user_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
        user_data.append((row[1],row[2])) # Skip the first column
        print (user_data)
    return user_data
def test_Login(driver,username,password):
    web_page_obj1=web_page(driver)
    web_page_obj1.open_page("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")
    web_page_obj1.login_page(username,password)
    assert "Dashboard" in driver.page_source ##validated with the sucessfully updated message

@pytest.mark.parametrize("username, password", read_excel_data())
def test_login_user_data(driver,username,password):
    test_Login(driver,username,password)


